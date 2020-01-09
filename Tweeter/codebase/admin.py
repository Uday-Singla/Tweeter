from django.contrib import admin
from django.http import HttpResponse
from .models import Profile, Post, Comment

class ProfileAdmin(admin.ModelAdmin):
    actions = ['export_xlsx', ]

    def export_xlsx(modeladmin, request, queryset):
        import openpyxl
        from openpyxl.utils import get_column_letter
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=profile.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.get_active_sheet()
        ws.title = "Profile"

        row_num = 0

        columns = [
            (u"User", 25),
            (u"Profile Picture", 150),
            (u"City", 20),
            (u"Date of Birth", 15),
        ]

        for col_num in range(len(columns)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = columns[col_num][0]
            # set column width
            ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

        for obj in queryset:
            row_num += 1
            row = [
                obj.user.username,
                obj.dp.url,
                obj.city,
                obj.dob,
            ]
            for col_num in range(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]

        wb.save(response)
        return response

    export_xlsx.short_description = u"Export XLSX"

admin.site.register(Profile, ProfileAdmin)
admin.site.register(Post)
admin.site.register(Comment)